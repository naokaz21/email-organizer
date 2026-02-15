"""
投資シミュレーションモジュール
収益シミュレーション（GPI→EGI→NOI→BTCFo）、投資判断ロジック、Excel出力
"""

import math
import numpy as np
import numpy_financial as npf
from io import BytesIO
from datetime import datetime

# ===== デフォルト投資パラメータ =====
DEFAULT_LTV = 0.90                    # 借入割合 90%
DEFAULT_INTEREST_RATE = 0.0225        # 金利 2.25%
DEFAULT_LOAN_TERM_YEARS = 30          # 返済期間 30年
DEFAULT_VACANCY_RATE = 0.05           # 空室率 5%
DEFAULT_RENT_DECLINE_RATE = 0.005     # 賃料下落率 0.5%/年
DEFAULT_HOLDING_PERIOD = 10           # 保有期間 10年
DEFAULT_EXPECTED_RETURN = 0.02        # 期待収益率（割引率）2.0%
DEFAULT_PURCHASE_EXPENSE_RATE = 0.08  # 購入諸費用 8%
DEFAULT_OPEX_RATIO = 0.15            # 運営費率 GPI比15%
DEFAULT_EXIT_CAP_RATE_SPREAD = 0.005  # 出口Cap Rate = 入口 + 0.5%
DEFAULT_SALE_EXPENSE_RATE = 0.04      # 売却諸費用 4%

# ===== 投資判断基準 =====
THRESHOLD_DCR = 1.2
THRESHOLD_BER = 0.80


def validate_simulation_inputs(data):
    """シミュレーション入力データのバリデーション"""
    warnings = []

    price = data.get('price')
    rent = data.get('full_occupancy_rent')

    if not price or price <= 0:
        return False, ["購入価格（price）が未設定または不正"]
    if not rent or rent <= 0:
        return False, ["満室想定賃料（full_occupancy_rent）が未設定または不正"]

    if not data.get('management_fee'):
        warnings.append("管理費が未抽出のためOPEXは概算率のみで計算")
    if not data.get('reserve_fund'):
        warnings.append("修繕積立金が未抽出のためOPEXは概算率のみで計算")
    if not data.get('total_units'):
        warnings.append("総戸数が未抽出")

    return True, warnings


def calculate_loan_payment(principal, annual_rate, years):
    """元利均等返済の月額返済額を計算"""
    if principal <= 0 or annual_rate <= 0 or years <= 0:
        return 0.0
    monthly_rate = annual_rate / 12
    n_payments = years * 12
    numerator = monthly_rate * (1 + monthly_rate) ** n_payments
    denominator = (1 + monthly_rate) ** n_payments - 1
    return principal * (numerator / denominator)


def calculate_remaining_balance(principal, annual_rate, total_months, paid_months):
    """ローン残高を計算"""
    if principal <= 0 or annual_rate <= 0:
        return 0.0
    r = annual_rate / 12
    balance = principal * ((1 + r) ** total_months - (1 + r) ** paid_months) / ((1 + r) ** total_months - 1)
    return max(balance, 0.0)


def build_annual_cashflows(params):
    """年次キャッシュフロー表を構築"""
    cashflows = []
    for year in range(1, params['holding_period'] + 1):
        # GPI: 賃料下落を反映
        gpi = params['full_occupancy_rent_annual'] * (1 - params['rent_decline_rate']) ** (year - 1)

        # EGI: 空室損控除
        vacancy_loss = gpi * params['vacancy_rate']
        egi = gpi - vacancy_loss

        # OPEX: GPI比率 + 抽出済み管理費・修繕積立金
        opex = gpi * params['opex_ratio']
        if params.get('management_fee_annual', 0) > 0:
            opex += params['management_fee_annual']
        if params.get('reserve_fund_annual', 0) > 0:
            opex += params['reserve_fund_annual']

        # NOI
        noi = egi - opex

        # BTCFo
        btcfo = noi - params['ads']

        cashflows.append({
            'year': year,
            'gpi': gpi,
            'vacancy_loss': vacancy_loss,
            'egi': egi,
            'opex': opex,
            'noi': noi,
            'ads': params['ads'],
            'btcfo': btcfo,
        })

    return cashflows


def calculate_sale_proceeds(noi_final_year, exit_cap_rate, loan_balance):
    """売却キャッシュフローを計算"""
    if exit_cap_rate <= 0:
        exit_cap_rate = 0.03  # 最低3%

    sale_price = noi_final_year / exit_cap_rate
    sale_expenses = sale_price * DEFAULT_SALE_EXPENSE_RATE
    net_proceeds = sale_price - sale_expenses - loan_balance

    return {
        'sale_price': sale_price,
        'sale_expenses': sale_expenses,
        'loan_balance': loan_balance,
        'net_proceeds': net_proceeds,
    }


def calculate_investment_metrics(cashflows, equity, sale, total_purchase_cost,
                                  loan_amount, ads, expected_return, purchase_price=None):
    """投資分析指標を計算"""
    year1 = cashflows[0]

    # FCR（総収益率）= 初年度NOI / 総投資額
    fcr = year1['noi'] / total_purchase_cost if total_purchase_cost > 0 else 0

    # K%（ローン定数）= ADS / 借入額
    k_percent = ads / loan_amount if loan_amount > 0 else 0

    # CCR（自己資本配当率）= 初年度BTCFo / 自己資金
    ccr = year1['btcfo'] / equity if equity > 0 else 0

    # DCR（借入償還余裕率）= 初年度NOI / ADS
    dcr = year1['noi'] / ads if ads > 0 else float('inf')

    # BER（損益分岐入居率）= (OPEX + ADS) / GPI
    ber = (year1['opex'] + ads) / year1['gpi'] if year1['gpi'] > 0 else 1.0

    # レバレッジ分析
    leverage = 'Positive' if fcr > k_percent else 'Negative'

    # IRR/NPV用キャッシュフロー配列
    cf_array = [-equity]
    for i, cf in enumerate(cashflows):
        if i == len(cashflows) - 1:
            cf_array.append(cf['btcfo'] + sale['net_proceeds'])
        else:
            cf_array.append(cf['btcfo'])

    # IRR
    try:
        irr = npf.irr(cf_array)
        if irr is None or np.isnan(irr):
            irr = None
    except Exception:
        irr = None

    # NPV
    try:
        npv = npf.npv(expected_return, cf_array)
        if npv is not None and np.isnan(npv):
            npv = None
    except Exception:
        npv = None

    # 表面利回り（参考値）= 満室想定年間賃料 / 物件価格
    # 一般的な表面利回りは購入諸費用を含まない物件価格で計算する
    price_for_yield = purchase_price if purchase_price and purchase_price > 0 else total_purchase_cost
    gross_yield = (cashflows[0]['gpi'] / price_for_yield) if price_for_yield > 0 else 0

    return {
        'gross_yield': gross_yield,
        'fcr': fcr,
        'k_percent': k_percent,
        'ccr': ccr,
        'dcr': dcr,
        'ber': ber,
        'leverage': leverage,
        'irr': irr,
        'npv': npv,
    }


def evaluate_investment_decision(metrics):
    """投資判断ロジック（6基準）"""
    decisions = {
        'fcr_vs_k': {
            'pass': metrics['fcr'] > metrics['k_percent'],
            'label': 'FCR > K%',
            'detail': f"FCR {metrics['fcr']:.2%} vs K% {metrics['k_percent']:.2%}",
        },
        'ccr_vs_fcr': {
            'pass': metrics['ccr'] > metrics['fcr'],
            'label': 'CCR > FCR',
            'detail': f"CCR {metrics['ccr']:.2%} vs FCR {metrics['fcr']:.2%}",
        },
        'dcr': {
            'pass': metrics['dcr'] >= THRESHOLD_DCR,
            'label': f'DCR >= {THRESHOLD_DCR}',
            'detail': f"DCR {metrics['dcr']:.2f}",
        },
        'ber': {
            'pass': metrics['ber'] <= THRESHOLD_BER,
            'label': f'BER <= {THRESHOLD_BER:.0%}',
            'detail': f"BER {metrics['ber']:.2%}",
        },
        'irr': {
            'pass': metrics['irr'] is not None and metrics['irr'] > DEFAULT_EXPECTED_RETURN,
            'label': f'IRR > {DEFAULT_EXPECTED_RETURN:.1%}',
            'detail': f"IRR {metrics['irr']:.2%}" if metrics['irr'] is not None else "IRR 計算不可",
        },
        'npv': {
            'pass': metrics['npv'] is not None and metrics['npv'] > 0,
            'label': 'NPV > 0',
            'detail': f"NPV {metrics['npv']:,.0f}円" if metrics['npv'] is not None else "NPV 計算不可",
        },
    }

    pass_count = sum(1 for d in decisions.values() if d['pass'])
    total_count = len(decisions)
    all_pass = pass_count == total_count

    return {
        'decisions': decisions,
        'all_pass': all_pass,
        'recommendation': '投資検討推奨' if all_pass else '投資見送り推奨',
        'pass_count': pass_count,
        'total_count': total_count,
    }


def run_simulation(comprehensive_data):
    """メインシミュレーション実行関数

    Args:
        comprehensive_data: extract_comprehensive_property_dataの返り値

    Returns:
        dict: シミュレーション結果全体。データ不足の場合はNone
    """
    try:
        # 1. バリデーション
        valid, warnings = validate_simulation_inputs(comprehensive_data)
        if not valid:
            print(f"シミュレーション入力不正: {warnings}")
            return None

        # 2. 基本パラメータ算出
        purchase_price = float(comprehensive_data['price'])
        purchase_expenses = purchase_price * DEFAULT_PURCHASE_EXPENSE_RATE
        total_purchase_cost = purchase_price + purchase_expenses
        loan_amount = purchase_price * DEFAULT_LTV
        equity = total_purchase_cost - loan_amount

        # 月額→年額変換
        full_occupancy_rent_monthly = float(comprehensive_data['full_occupancy_rent'])
        full_occupancy_rent_annual = full_occupancy_rent_monthly * 12

        management_fee_monthly = float(comprehensive_data.get('management_fee') or 0)
        reserve_fund_monthly = float(comprehensive_data.get('reserve_fund') or 0)

        # ローン計算
        monthly_payment = calculate_loan_payment(loan_amount, DEFAULT_INTEREST_RATE, DEFAULT_LOAN_TERM_YEARS)
        ads = monthly_payment * 12

        # 3. 年次CF構築パラメータ
        params = {
            'full_occupancy_rent_annual': full_occupancy_rent_annual,
            'vacancy_rate': DEFAULT_VACANCY_RATE,
            'rent_decline_rate': DEFAULT_RENT_DECLINE_RATE,
            'opex_ratio': DEFAULT_OPEX_RATIO,
            'management_fee_annual': management_fee_monthly * 12,
            'reserve_fund_annual': reserve_fund_monthly * 12,
            'ads': ads,
            'holding_period': DEFAULT_HOLDING_PERIOD,
        }

        cashflows = build_annual_cashflows(params)

        # 4. 売却CF計算
        entry_cap_rate = cashflows[0]['noi'] / total_purchase_cost if total_purchase_cost > 0 else 0.05
        exit_cap_rate = max(entry_cap_rate + DEFAULT_EXIT_CAP_RATE_SPREAD, 0.03)

        total_months = DEFAULT_LOAN_TERM_YEARS * 12
        paid_months = DEFAULT_HOLDING_PERIOD * 12
        loan_balance = calculate_remaining_balance(loan_amount, DEFAULT_INTEREST_RATE, total_months, paid_months)

        sale = calculate_sale_proceeds(cashflows[-1]['noi'], exit_cap_rate, loan_balance)

        # 5. 投資指標計算
        metrics = calculate_investment_metrics(
            cashflows, equity, sale, total_purchase_cost,
            loan_amount, ads, DEFAULT_EXPECTED_RETURN,
            purchase_price=purchase_price
        )

        # 6. 投資判断
        decision = evaluate_investment_decision(metrics)

        return {
            'params': {
                'purchase_price': purchase_price,
                'purchase_expenses': purchase_expenses,
                'total_purchase_cost': total_purchase_cost,
                'loan_amount': loan_amount,
                'equity': equity,
                'ltv': DEFAULT_LTV,
                'interest_rate': DEFAULT_INTEREST_RATE,
                'loan_term': DEFAULT_LOAN_TERM_YEARS,
                'monthly_payment': monthly_payment,
                'ads': ads,
                'vacancy_rate': DEFAULT_VACANCY_RATE,
                'rent_decline_rate': DEFAULT_RENT_DECLINE_RATE,
                'holding_period': DEFAULT_HOLDING_PERIOD,
                'opex_ratio': DEFAULT_OPEX_RATIO,
                'exit_cap_rate': exit_cap_rate,
                'expected_return': DEFAULT_EXPECTED_RETURN,
                'full_occupancy_rent_monthly': full_occupancy_rent_monthly,
                'full_occupancy_rent_annual': full_occupancy_rent_annual,
                'management_fee_monthly': management_fee_monthly,
                'reserve_fund_monthly': reserve_fund_monthly,
            },
            'cashflows': cashflows,
            'sale': sale,
            'metrics': metrics,
            'decision': decision,
            'warnings': warnings,
        }

    except Exception as e:
        print(f"シミュレーション実行エラー: {e}")
        import traceback
        traceback.print_exc()
        return None


def create_simulation_excel(simulation_result, property_info, drive_service, folder_id):
    """Excelファイルを作成してGoogle Driveに保存（数式ベース）

    Args:
        simulation_result: run_simulationの返り値
        property_info: {"property_number": str, "station": str}
        drive_service: Google Drive APIサービス
        folder_id: 保存先フォルダID

    Returns:
        str: アップロードしたファイルのID。失敗時はNone
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        p = simulation_result['params']
        m = simulation_result['metrics']
        d = simulation_result['decision']
        cfs = simulation_result['cashflows']
        sale = simulation_result['sale']

        # スタイル定義
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        section_font = Font(bold=True, size=12)
        yen_format = '#,##0'
        pct_format = '0.00%'
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def styled_cell(ws, r, c, value, fmt=None, font=None, fill=None, border=True):
            cell = ws.cell(row=r, column=c, value=value)
            if fmt:
                cell.number_format = fmt
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = thin_border
            return cell

        # ===== Sheet 1: 入力パラメータ（黄色セル=変更可） =====
        ws1 = wb.active
        ws1.title = "入力パラメータ"
        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 22
        ws1.column_dimensions['C'].width = 12

        row = 1
        ws1.cell(row=row, column=1, value="投資シミュレーション").font = title_font
        row += 1
        ws1.cell(row=row, column=1, value=f"物件番号: {property_info.get('property_number', '')}  駅: {property_info.get('station', '')}")
        row += 1
        ws1.cell(row=row, column=1, value=f"作成日: {datetime.now().strftime('%Y-%m-%d')}")
        row += 1
        ws1.cell(row=row, column=1, value="※ 黄色セルは変更可能（数式が自動再計算されます）").font = Font(italic=True, size=9, color="666666")
        row += 2

        # 入力パラメータセクション - 黄色=ユーザー変更可
        ws1.cell(row=row, column=1, value="基本入力パラメータ").font = section_font
        row += 1
        # 名前付きセル位置を記録（数式参照用）
        cells = {}
        input_params = [
            ("purchase_price", "購入価格", p['purchase_price'], yen_format, True),
            ("expense_rate", "購入諸費用率", DEFAULT_PURCHASE_EXPENSE_RATE, pct_format, True),
            ("ltv", "LTV（借入割合）", p['ltv'], pct_format, True),
            ("interest_rate", "ローン金利", p['interest_rate'], pct_format, True),
            ("loan_term", "返済期間（年）", p['loan_term'], '#,##0', True),
            ("vacancy_rate", "空室率", p['vacancy_rate'], pct_format, True),
            ("rent_decline_rate", "賃料下落率（年率）", p['rent_decline_rate'], pct_format, True),
            ("holding_period", "保有期間（年）", p['holding_period'], '#,##0', True),
            ("expected_return", "期待収益率（割引率）", p['expected_return'], pct_format, True),
            ("opex_ratio", "運営費率（GPI比）", p['opex_ratio'], pct_format, True),
            ("exit_cap_spread", "出口Cap Rateスプレッド", DEFAULT_EXIT_CAP_RATE_SPREAD, pct_format, True),
            ("sale_expense_rate", "売却諸費用率", DEFAULT_SALE_EXPENSE_RATE, pct_format, True),
            ("rent_monthly", "満室想定賃料（月額）", p['full_occupancy_rent_monthly'], yen_format, True),
            ("mgmt_fee_monthly", "管理費（月額）", p['management_fee_monthly'], yen_format, True),
            ("reserve_fund_monthly", "修繕積立金（月額）", p['reserve_fund_monthly'], yen_format, True),
        ]
        for key, label, value, fmt, is_input in input_params:
            styled_cell(ws1, row, 1, label)
            c = styled_cell(ws1, row, 2, value, fmt=fmt, fill=input_fill if is_input else None)
            cells[key] = f'B{row}'
            row += 1

        row += 1
        ws1.cell(row=row, column=1, value="算出パラメータ（数式）").font = section_font
        row += 1

        # 算出パラメータ = 数式セル
        # 購入諸費用
        styled_cell(ws1, row, 1, "購入諸費用")
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        ws1.cell(row=row, column=2).value = f"={cells['purchase_price']}*{cells['expense_rate']}"
        cells['purchase_expenses'] = f'B{row}'
        row += 1

        # 総投資額
        styled_cell(ws1, row, 1, "総投資額")
        ws1.cell(row=row, column=2).value = f"={cells['purchase_price']}+{cells['purchase_expenses']}"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['total_cost'] = f'B{row}'
        row += 1

        # 借入額
        styled_cell(ws1, row, 1, "借入額")
        ws1.cell(row=row, column=2).value = f"={cells['purchase_price']}*{cells['ltv']}"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['loan_amount'] = f'B{row}'
        row += 1

        # 自己資金
        styled_cell(ws1, row, 1, "自己資金")
        ws1.cell(row=row, column=2).value = f"={cells['total_cost']}-{cells['loan_amount']}"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['equity'] = f'B{row}'
        row += 1

        # 月額返済額 (PMT関数)
        styled_cell(ws1, row, 1, "月額返済額")
        ws1.cell(row=row, column=2).value = f"=-PMT({cells['interest_rate']}/12,{cells['loan_term']}*12,{cells['loan_amount']})"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['monthly_payment'] = f'B{row}'
        row += 1

        # ADS
        styled_cell(ws1, row, 1, "年間返済額（ADS）")
        ws1.cell(row=row, column=2).value = f"={cells['monthly_payment']}*12"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['ads'] = f'B{row}'
        row += 1

        # 満室想定賃料（年額）
        styled_cell(ws1, row, 1, "満室想定賃料（年額）")
        ws1.cell(row=row, column=2).value = f"={cells['rent_monthly']}*12"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['rent_annual'] = f'B{row}'
        row += 1

        # 管理費（年額）
        styled_cell(ws1, row, 1, "管理費（年額）")
        ws1.cell(row=row, column=2).value = f"={cells['mgmt_fee_monthly']}*12"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['mgmt_annual'] = f'B{row}'
        row += 1

        # 修繕積立金（年額）
        styled_cell(ws1, row, 1, "修繕積立金（年額）")
        ws1.cell(row=row, column=2).value = f"={cells['reserve_fund_monthly']}*12"
        styled_cell(ws1, row, 2, None, fmt=yen_format)
        cells['reserve_annual'] = f'B{row}'
        row += 1

        # ===== Sheet 2: 年次キャッシュフロー（数式） =====
        ws2 = wb.create_sheet("年次キャッシュフロー")
        ref = "'入力パラメータ'"  # シート参照

        headers = ["年度", "GPI（満室想定）", "空室損", "EGI（実効収入）", "OPEX（運営費）", "NOI（営業純利益）", "ADS（返済額）", "BTCFo（税引前CF）"]
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        ws2.column_dimensions['A'].width = 8
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws2.column_dimensions[col_letter].width = 20

        even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        holding = p['holding_period']

        for i in range(holding):
            r = i + 2
            year = i + 1
            # A: 年度
            styled_cell(ws2, r, 1, year)
            # B: GPI = 満室年額 * (1 - 賃料下落率)^(年-1)
            ws2.cell(row=r, column=2).value = f"={ref}!{cells['rent_annual']}*(1-{ref}!{cells['rent_decline_rate']})^(A{r}-1)"
            # C: 空室損 = GPI * 空室率
            ws2.cell(row=r, column=3).value = f"=B{r}*{ref}!{cells['vacancy_rate']}"
            # D: EGI = GPI - 空室損
            ws2.cell(row=r, column=4).value = f"=B{r}-C{r}"
            # E: OPEX = GPI * 運営費率 + 管理費年額 + 修繕積立金年額
            ws2.cell(row=r, column=5).value = f"=B{r}*{ref}!{cells['opex_ratio']}+{ref}!{cells['mgmt_annual']}+{ref}!{cells['reserve_annual']}"
            # F: NOI = EGI - OPEX
            ws2.cell(row=r, column=6).value = f"=D{r}-E{r}"
            # G: ADS
            ws2.cell(row=r, column=7).value = f"={ref}!{cells['ads']}"
            # H: BTCFo = NOI - ADS
            ws2.cell(row=r, column=8).value = f"=F{r}-G{r}"

            for col in range(2, 9):
                c = ws2.cell(row=r, column=col)
                c.number_format = yen_format
                c.border = thin_border
                if i % 2 == 1:
                    c.fill = even_fill
            ws2.cell(row=r, column=1).border = thin_border
            if i % 2 == 1:
                ws2.cell(row=r, column=1).fill = even_fill

        # 合計行
        total_row = holding + 2
        styled_cell(ws2, total_row, 1, "合計", font=Font(bold=True))
        for col in range(2, 9):
            col_letter = get_column_letter(col)
            ws2.cell(row=total_row, column=col).value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
            c = ws2.cell(row=total_row, column=col)
            c.number_format = yen_format
            c.font = Font(bold=True)
            c.border = thin_border

        ws2.freeze_panes = 'A2'

        # ===== Sheet 3: 売却シミュレーション（数式） =====
        ws3 = wb.create_sheet("売却シミュレーション")
        ws3.column_dimensions['A'].width = 28
        ws3.column_dimensions['B'].width = 22

        cf_sheet = "'年次キャッシュフロー'"
        last_cf_row = holding + 1  # 最終年度のデータ行

        row = 1
        ws3.cell(row=row, column=1, value="売却シミュレーション").font = title_font
        row += 2

        # 入口Cap Rate（NOI初年度/総投資額）
        styled_cell(ws3, row, 1, "入口Cap Rate")
        ws3.cell(row=row, column=2).value = f"={cf_sheet}!F2/{ref}!{cells['total_cost']}"
        ws3.cell(row=row, column=2).number_format = pct_format
        cells['entry_cap'] = f'B{row}'
        row += 1

        # 出口Cap Rate
        styled_cell(ws3, row, 1, "出口Cap Rate")
        ws3.cell(row=row, column=2).value = f"=MAX({cells['entry_cap']}+{ref}!{cells['exit_cap_spread']},0.03)"
        ws3.cell(row=row, column=2).number_format = pct_format
        cells['exit_cap'] = f'B{row}'
        row += 1

        # 最終年NOI
        styled_cell(ws3, row, 1, "最終年NOI")
        ws3.cell(row=row, column=2).value = f"={cf_sheet}!F{last_cf_row}"
        ws3.cell(row=row, column=2).number_format = yen_format
        cells['final_noi'] = f'B{row}'
        row += 1

        # 売却想定価格 = NOI/Cap Rate
        styled_cell(ws3, row, 1, "売却想定価格")
        ws3.cell(row=row, column=2).value = f"={cells['final_noi']}/{cells['exit_cap']}"
        ws3.cell(row=row, column=2).number_format = yen_format
        cells['sale_price'] = f'B{row}'
        row += 1

        # 売却諸費用
        styled_cell(ws3, row, 1, "売却諸費用")
        ws3.cell(row=row, column=2).value = f"={cells['sale_price']}*{ref}!{cells['sale_expense_rate']}"
        ws3.cell(row=row, column=2).number_format = yen_format
        cells['sale_expenses'] = f'B{row}'
        row += 1

        # ローン残高 (FV関数で計算)
        styled_cell(ws3, row, 1, "ローン残高")
        # FV(rate, nper, pmt, pv) で保有期間後の残高を計算
        ws3.cell(row=row, column=2).value = (
            f"=-FV({ref}!{cells['interest_rate']}/12,"
            f"{ref}!{cells['holding_period']}*12,"
            f"{ref}!{cells['monthly_payment']},"
            f"-{ref}!{cells['loan_amount']})"
        )
        ws3.cell(row=row, column=2).number_format = yen_format
        cells['loan_balance'] = f'B{row}'
        row += 1

        # 売却手取り
        styled_cell(ws3, row, 1, "売却手取り（税引前）")
        ws3.cell(row=row, column=2).value = f"={cells['sale_price']}-{cells['sale_expenses']}-{cells['loan_balance']}"
        ws3.cell(row=row, column=2).number_format = yen_format
        ws3.cell(row=row, column=2).font = Font(bold=True)
        cells['net_proceeds'] = f'B{row}'
        row += 2

        # ===== IRR計算用CF =====
        ws3.cell(row=row, column=1, value="IRR計算用キャッシュフロー").font = section_font
        row += 1
        for col, h in enumerate(["年度", "キャッシュフロー"], 1):
            c = ws3.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        row += 1

        irr_start_row = row
        # 初期投資 (年度0)
        styled_cell(ws3, row, 1, "0（初期投資）")
        ws3.cell(row=row, column=2).value = f"=-{ref}!{cells['equity']}"
        ws3.cell(row=row, column=2).number_format = yen_format
        ws3.cell(row=row, column=2).border = thin_border
        row += 1

        # 各年CF
        for i in range(holding):
            cf_r = i + 2  # 年次CFシートのデータ行
            styled_cell(ws3, row, 1, i + 1)
            if i == holding - 1:
                # 最終年: BTCFo + 売却手取り
                ws3.cell(row=row, column=2).value = f"={cf_sheet}!H{cf_r}+{cells['net_proceeds']}"
            else:
                ws3.cell(row=row, column=2).value = f"={cf_sheet}!H{cf_r}"
            ws3.cell(row=row, column=2).number_format = yen_format
            ws3.cell(row=row, column=2).border = thin_border
            row += 1

        irr_end_row = row - 1

        row += 1
        styled_cell(ws3, row, 1, "IRR", font=Font(bold=True))
        ws3.cell(row=row, column=2).value = f"=IFERROR(IRR(B{irr_start_row}:B{irr_end_row}),\"計算不可\")"
        ws3.cell(row=row, column=2).number_format = pct_format
        ws3.cell(row=row, column=2).font = Font(bold=True)
        ws3.cell(row=row, column=2).border = thin_border
        cells['irr'] = f'B{row}'
        row += 1

        styled_cell(ws3, row, 1, "NPV", font=Font(bold=True))
        ws3.cell(row=row, column=2).value = f"=IFERROR(NPV({ref}!{cells['expected_return']},B{irr_start_row+1}:B{irr_end_row})+B{irr_start_row},\"計算不可\")"
        ws3.cell(row=row, column=2).number_format = yen_format
        ws3.cell(row=row, column=2).font = Font(bold=True)
        ws3.cell(row=row, column=2).border = thin_border
        cells['npv'] = f'B{row}'

        # ===== Sheet 4: 投資指標（数式） =====
        ws4 = wb.create_sheet("投資指標")
        ws4.column_dimensions['A'].width = 28
        ws4.column_dimensions['B'].width = 18
        ws4.column_dimensions['C'].width = 22

        row = 1
        ws4.cell(row=row, column=1, value="投資指標サマリー").font = title_font
        row += 2

        for col, h in enumerate(["指標", "算出値", "判定"], 1):
            c = ws4.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        row += 1

        sale_ref = "'売却シミュレーション'"

        # 表面利回り = 満室年額 / 購入価格
        styled_cell(ws4, row, 1, "表面利回り")
        ws4.cell(row=row, column=2).value = f"={ref}!{cells['rent_annual']}/{ref}!{cells['purchase_price']}"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        styled_cell(ws4, row, 3, "参考値")
        cells['gross_yield'] = f'B{row}'
        row += 1

        # FCR = 初年度NOI / 総投資額
        styled_cell(ws4, row, 1, "FCR（総収益率）")
        ws4.cell(row=row, column=2).value = f"={cf_sheet}!F2/{ref}!{cells['total_cost']}"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        cells['fcr'] = f'B{row}'
        ws4.cell(row=row, column=3).value = f'=IF({cells["fcr"]}>{cells["k_pct"]},"○ FCR > K%","× FCR ≤ K%")' if False else None
        fcr_row = row
        row += 1

        # K%
        styled_cell(ws4, row, 1, "K%（ローン定数）")
        ws4.cell(row=row, column=2).value = f"={ref}!{cells['ads']}/{ref}!{cells['loan_amount']}"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        cells['k_pct'] = f'B{row}'
        k_row = row
        row += 1

        # 判定: FCR > K%
        ws4.cell(row=fcr_row, column=3).value = f'=IF(B{fcr_row}>B{k_row},"○ FCR > K%","× FCR ≤ K%")'
        ws4.cell(row=fcr_row, column=3).border = thin_border
        styled_cell(ws4, k_row, 3, "参考値")

        # CCR
        styled_cell(ws4, row, 1, "CCR（自己資本配当率）")
        ws4.cell(row=row, column=2).value = f"={cf_sheet}!H2/{ref}!{cells['equity']}"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=3).value = f'=IF(B{row}>B{fcr_row},"○ CCR > FCR","× CCR ≤ FCR")'
        ws4.cell(row=row, column=3).border = thin_border
        cells['ccr'] = f'B{row}'
        row += 1

        # レバレッジ
        styled_cell(ws4, row, 1, "レバレッジ分析")
        ws4.cell(row=row, column=2).value = f'=IF(B{fcr_row}>B{k_row},"Positive","Negative")'
        ws4.cell(row=row, column=2).border = thin_border
        styled_cell(ws4, row, 3, "")
        row += 1

        # DCR
        styled_cell(ws4, row, 1, "DCR（借入償還余裕率）")
        ws4.cell(row=row, column=2).value = f"={cf_sheet}!F2/{ref}!{cells['ads']}"
        ws4.cell(row=row, column=2).number_format = '0.00'
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=3).value = f'=IF(B{row}>=1.2,"○ DCR ≥ 1.2","× DCR < 1.2")'
        ws4.cell(row=row, column=3).border = thin_border
        row += 1

        # BER
        styled_cell(ws4, row, 1, "BER（損益分岐入居率）")
        ws4.cell(row=row, column=2).value = f"=({cf_sheet}!E2+{ref}!{cells['ads']})/{cf_sheet}!B2"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=3).value = f'=IF(B{row}<=0.8,"○ BER ≤ 80%","× BER > 80%")'
        ws4.cell(row=row, column=3).border = thin_border
        row += 1

        # IRR
        styled_cell(ws4, row, 1, "IRR（内部収益率）")
        ws4.cell(row=row, column=2).value = f"={sale_ref}!{cells['irr']}"
        ws4.cell(row=row, column=2).number_format = pct_format
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=3).value = f'=IF(ISNUMBER(B{row}),IF(B{row}>{ref}!{cells["expected_return"]},"○ IRR > 期待収益率","× IRR ≤ 期待収益率"),"計算不可")'
        ws4.cell(row=row, column=3).border = thin_border
        row += 1

        # NPV
        styled_cell(ws4, row, 1, "NPV（正味現在価値）")
        ws4.cell(row=row, column=2).value = f"={sale_ref}!{cells['npv']}"
        ws4.cell(row=row, column=2).number_format = yen_format
        ws4.cell(row=row, column=2).border = thin_border
        ws4.cell(row=row, column=3).value = f'=IF(ISNUMBER(B{row}),IF(B{row}>0,"○ NPV > 0","× NPV ≤ 0"),"計算不可")'
        ws4.cell(row=row, column=3).border = thin_border
        row += 2

        # 総合判定
        ws4.cell(row=row, column=1, value="総合判定").font = section_font
        ws4.cell(row=row, column=1).border = thin_border
        judge = ws4.cell(row=row, column=2, value=f"{d['recommendation']}（{d['pass_count']}/{d['total_count']}項目クリア）")
        judge.font = Font(bold=True, size=12)
        judge.fill = pass_fill if d['all_pass'] else fail_fill
        judge.border = thin_border

        # Excel保存 → Google Drive アップロード
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from googleapiclient.http import MediaIoBaseUpload
        filename = f"投資シミュレーション_{property_info.get('property_number', 'unknown')}_{property_info.get('station', '')}.xlsx"
        media = MediaIoBaseUpload(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        uploaded = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        print(f"Excel保存完了: {filename}")
        return uploaded['id']

    except Exception as e:
        print(f"Excel作成エラー: {e}")
        import traceback
        traceback.print_exc()
        return None


def format_simulation_summary_for_report(simulation_result):
    """Google Docsレポート用のシミュレーションサマリーテキストを生成"""
    if not simulation_result:
        return ["", "【投資シミュレーション】", "シミュレーション実行不可（データ不足）"]

    p = simulation_result['params']
    m = simulation_result['metrics']
    d = simulation_result['decision']

    def mark(passed):
        return "○" if passed else "×"

    lines = [
        "",
        "【投資シミュレーション結果】",
        f"総投資額: {p['total_purchase_cost']:,.0f}円（購入価格 {p['purchase_price']:,.0f}円 + 諸費用 {p['purchase_expenses']:,.0f}円）",
        f"借入: {p['loan_amount']:,.0f}円（LTV {p['ltv']:.0%}）/ 自己資金: {p['equity']:,.0f}円",
        f"金利: {p['interest_rate']:.2%} / 期間: {p['loan_term']}年 / ADS: {p['ads']:,.0f}円/年",
        "",
        "【投資指標】",
        f"表面利回り: {m['gross_yield']:.2%}（参考値）",
        f"FCR（総収益率）: {m['fcr']:.2%}  {mark(d['decisions']['fcr_vs_k']['pass'])}",
        f"K%（ローン定数）: {m['k_percent']:.2%}",
        f"CCR（自己資本配当率）: {m['ccr']:.2%}  {mark(d['decisions']['ccr_vs_fcr']['pass'])}",
        f"レバレッジ分析: {m['leverage']}",
        f"DCR（借入償還余裕率）: {m['dcr']:.2f}  {mark(d['decisions']['dcr']['pass'])}",
        f"BER（損益分岐入居率）: {m['ber']:.2%}  {mark(d['decisions']['ber']['pass'])}",
    ]

    if m.get('irr') is not None:
        lines.append(f"IRR（内部収益率）: {m['irr']:.2%}  {mark(d['decisions']['irr']['pass'])}")
    else:
        lines.append("IRR（内部収益率）: 計算不可  ×")

    if m.get('npv') is not None:
        lines.append(f"NPV（正味現在価値）: {m['npv']:,.0f}円  {mark(d['decisions']['npv']['pass'])}")
    else:
        lines.append("NPV（正味現在価値）: 計算不可  ×")

    lines.extend([
        "",
        f"総合判定: {d['recommendation']}（{d['pass_count']}/{d['total_count']}項目クリア）",
    ])

    # 警告があれば追加
    if simulation_result.get('warnings'):
        lines.append("")
        lines.append("※ 注意事項:")
        for w in simulation_result['warnings']:
            lines.append(f"  - {w}")

    return lines
